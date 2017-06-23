#!/usr/bin/python3

''' This module reads the limits from the selected excel limits file and creates a dictionary
containing the lower/upper limits for each temp/voltage/mode condition. This limit dictionary 
is used by the other modules for limit anlysis. '''

#### LIMITS ####

from openpyxl import Workbook, load_workbook
import pprint

class Limits(object):
    def __init__(self, filepath, sheet, boards, temps, binning=False):
        self.filepath = filepath
        self.sheet = sheet
        self.wb = load_workbook(self.filepath, read_only=True)
        self.ws = self.wb[self.sheet]
        self.boards = boards  # e.g. - 'B1' or 'B4'
        self.temps = temps  # integers, e.g. - 85 or -40 or 23
        self.binning = binning # boolean for if LED binning used for project limits file

        self.lim = {} # holds current limits
        self.modules = []  # e.g. - 'PARK' or 'TURN'
        self.board_modes = []  # e.g. - 'B3B4' or 'B2'
        self.module_modes = []  # e.g. - 'DRLTURN' or 'PARK'
        self.board_module_pairs = {}  # keys: boards; values: modules
        self.temp_rows = {}  # row reference for each temperature
        self.mode_row = ''   # row in which mode headers are located
        self.mode_cols = {}  # column for each board_mode key
        self.outage_link_board = ''
        self.outage_present = False
        self.voltage_column = 2
        self.voltage_header_row = 13
        self.voltages = []
        self.b1_row = None
        self.led_binning_dict = {}

        self.__get_boards()
        self.__get_temp_rows()
        self.__get_mode_row()
        self.__get_modes_and_mode_cols()
        self.__get_voltage_header()
        self.__get_voltages()
        self.fill_current_limits()
        self.fill_outage_limits()

    def __get_boards(self):
        self.get_b1_row()
        i = 0
        for board in self.boards:
            self.board_module_pairs[board] = self.ws.cell(row = self.b1_row+i, column = 2).value
            link = self.ws.cell(row = self.b1_row+i, column = 3).value
            if link == 'Y':
                self.outage_link_board = board
                self.outage_present = True
            led_bins = self.ws.cell(row = self.b1_row+i, column = 4).value
            if led_bins:
                self.led_binning_dict[board] = led_bins.split(' ')
            i += 1
        self.modules = [self.board_module_pairs[board] for board in self.boards]

    def __get_temp_rows(self):
        cell_range = self.ws.get_squared_range(min_col=1, min_row=1, max_col=1, max_row=self.ws.max_row)
        for row in cell_range:
            for cell in row:
                for temp in self.temps:
                    if (str(temp)+'C') == cell.value:
                        self.temp_rows[temp] = cell.row

    def __get_mode_row(self):
        cell_range = self.ws.get_squared_range(min_col=1, min_row=1, max_col=1, max_row=self.ws.max_row)
        for row in cell_range:
            for cell in row:
                value_string = str(cell.value).lower()
                if cell.value and (value_string == 'modes') or (value_string == 'mode'):
                    self.mode_row = cell.row

    def __get_modes_and_mode_cols(self):
        ''' Finds all modes and their column numbers, including variable led binning ''' 
        cell_range = self.ws.get_squared_range(min_col=1, min_row=self.mode_row, 
                                               max_col=self.ws.max_column, max_row=self.mode_row)
        for row in cell_range:
            for cell in row:
                if cell.column and cell.value and any(module in cell.value for module in self.modules):
                    module_mode = cell.value
                    self.module_modes.append(module_mode)
                    #board_mode = self.translate_module_mode_to_board_mode(cell.value)
                    #self.board_modes.append(board_mode)
                    self.mode_cols[module_mode] = cell.column

    def __get_voltage_header(self):
        ''' Info '''
        cell_range = self.ws.get_squared_range(min_col=self.voltage_column, min_row=1, 
                                               max_col=self.voltage_column, max_row=self.ws.max_row)
        for row in cell_range:
            for cell in row:
                value_string = str(cell.value).lower()
                if cell.value and (value_string == 'voltages') or (value_string == 'voltage'):
                    self.voltage_header_row = cell.row

    def __get_voltages(self):
        cell_range = self.ws.get_squared_range(min_col=self.voltage_column, min_row=self.voltage_header_row+1, 
                                               max_col=self.voltage_column, max_row=self.voltage_header_row+10)
        for row in cell_range:
            for cell in row:
                if cell.value and (cell.value not in self.voltages):
                    self.voltages.append(cell.value)


    def translate_module_mode_to_board_mode(self, module_mode):
        ''' Input: e.g. - 'DRLTURN' or 'LB' 
            Output: e.g. - 'B3B4' or 'B2' '''
        board_mode = module_mode
        for board, module in self.board_module_pairs.items():
            board_mode = board_mode.replace(module, board, 1) # replace only first instance
        return board_mode

    def get_b1_row(self):
        ''' Finds the row of the 'B1' label '''
        cell_range = self.ws.get_squared_range(min_col=1, min_row=1, max_col=1, max_row=11)
        for row in cell_range:
            for cell in row:
                if cell.value and cell.value.lower() == 'b1':
                    self.b1_row = cell.row

    def fill_current_limits(self):
        if self.led_binning_dict:
            self.fill_limits_binning()
        else:
            self.fill_limits_no_binning()

    def fill_limits_no_binning(self):
        ''' Build limits dictionary from limits file '''
        for temp in self.temps:
            temp_row = self.temp_rows[temp]
            self.lim[temp] = {}
            for board_mode, column in self.mode_cols.items():
                self.lim[temp][board_mode] = {}
                for i in range(len(self.voltages)):
                    voltage = float(self.ws.cell(row=temp_row+i, column=self.voltage_column).value)
                    lim_min = self.ws.cell(row=temp_row+i, column=column).value
                    lim_max = self.ws.cell(row=temp_row+i, column=column+1).value
                    lim_min = round(float(lim_min), 3)
                    lim_max = round(float(lim_max), 3)
                    self.lim[temp][board_mode][voltage] = (lim_min, lim_max)

    def fill_limits_binning(self):
        ''' Build limits dictionary from limits file '''
        for temp in self.temps:
            temp_row = self.temp_rows[temp]
            self.lim[temp] = {}
            for module_mode, column in self.mode_cols.items():
                self.lim[temp][module_mode] = {}
                for i in range(len(self.voltages)):
                    voltage = float(self.ws.cell(row=temp_row+i, column=self.voltage_column).value)
                    lim_min = self.ws.cell(row=temp_row+i, column=column).value
                    lim_max = self.ws.cell(row=temp_row+i, column=column+1).value
                    lim_min = round(float(lim_min), 3)
                    lim_max = round(float(lim_max), 3)
                    self.lim[temp][module_mode][voltage] = (lim_min, lim_max)

    def fill_outage_limits(self):
        if self.outage_present:
            outage = self.board_module_pairs['B6']
            self.lim[outage] = {}
            ## OFF limits
            off_min = self.ws.cell(row=5, column=7).value
            off_max = self.ws.cell(row=5, column=8).value
            self.lim[outage]['OFF'] = (off_min, off_max)
            ## ON limits
            i = 0
            self.lim[outage]['ON'] = {}
            for voltage in self.voltages:
                on_min = self.ws.cell(row=8+i, column=7).value
                on_max = self.ws.cell(row=8+i, column=8).value
                self.lim[outage]['ON'][voltage] = (on_min, on_max)
                i+=1

    def print_info(self):
        print('Board/Module Pairs:', self.board_module_pairs)
        print('\nMode Columns:', self.mode_cols, '\n')
        pp = pprint.PrettyPrinter()
        pp.pprint(self.lim)
        print('\n\n')



#!/usr/bin/python3

''' This module reads the limits from the selected excel limits file and creates a dictionary
containing the lower/upper limits for each temp/voltage/mode condition. This limit dictionary 
is used by the other modules for limit anlysis. '''

#### LIMITS ####

from openpyxl import Workbook, load_workbook
import pprint

class Limits(object):
    def __init__(self, filepath, sheet, boards, temps):
        self.filepath = filepath
        self.sheet = sheet
        self.wb = load_workbook(self.filepath, read_only=True)
        self.ws = self.wb[self.sheet]
        self.boards = boards  # e.g. - 'B1' or 'B4'
        self.modules = []  # e.g. - 'PARK' or 'TURN'
        self.board_modes = []  # e.g. - 'B3B4' or 'B2'
        self.module_modes = []  # e.g. - 'DRLTURN' or 'PARK'
        self.board_module_pairs = {}  # keys: boards; values: modules
        self.temps = temps  # integers, e.g. - 85 or -40 or 23
        self.temp_rows = {}  # row reference for each temperature
        self.mode_row = ''   # row in which mode headers are located
        self.mode_cols = {}  # column for each mode header
        self.outage_link_board = '' 

        self.__get_boards()
        self.__get_temp_rows()
        self.__get_mode_row()
        self.__get_modes_and_mode_cols()


    def __get_boards(self):
        self.get_b1_row()
        i = 0
        for board in self.boards:
            self.board_module_pairs[board] = self.ws.cell(row = self.b1_row+i, column = 2).value
            link = self.ws.cell(row = self.b1_row+i, column = 3).value
            if link == 'Y':
                self.outage_link_board = board
            i += 1
        self.modules = [self.board_module_pairs[board] for board in self.boards]

    def __get_temp_rows(self):
        cell_range = self.ws.get_squared_range(min_col=1, min_row=1, max_col=1, max_row=100)
        for row in cell_range:
            for cell in row:
                for temp in self.temps:
                    if (str(temp)+'C') == cell.value:
                        self.temp_rows[temp] = cell.row

    def __get_mode_row(self):
        cell_range = self.ws.get_squared_range(min_col=1, min_row=1, max_col=1, max_row=100)
        for row in cell_range:
            for cell in row:
                value_string = str(cell.value).lower()
                if cell.value and (value_string == 'modes') or (value_string == 'mode'):
                    self.mode_row = cell.row

    def __get_modes_and_mode_cols(self):
        ''' Finds all modes and their column numbers ''' 
        cell_range = self.ws.get_squared_range(min_col=1, min_row=self.mode_row, max_col=100, max_row=self.mode_row)
        for row in cell_range:
            for cell in row:
                if cell.column and cell.value and any(module in cell.value for module in self.modules):
                    self.module_modes.append(cell.value)
                    board_mode = self.translate_module_mode_to_board_mode(cell.value)
                    self.mode_cols[board_mode] = cell.column

    def translate_module_mode_to_board_mode(self, module_mode):
        ''' Input: e.g. - 'DRLTURN' or 'LB' 
            Output: e.g. - 'B3B4' or 'B2' '''
        board_mode = module_mode
        for board, module in self.board_module_pairs.items():
            board_mode = board_mode.replace(module, board, 1) # replace only first instance
        return board_mode

    def get_b1_row(self):
        ''' Finds the row of the first board label '''
        cell_range = self.ws.get_squared_range(min_col=1, min_row=1, max_col=1, max_row=11)
        for row in cell_range:
            for cell in row:
                if cell.value and cell.value.lower() == 'b1':
                    self.b1_row = cell.row

    # def fill_limits(self):
    #     wb = load_workbook(self.filepath, read_only=True)
    #     ws = wb[self.sheet]
    #     temps = {23:self.row23, -40:self.rowm40, 85:self.row85, 45:self.row45, 60:self.row60, 70:self.row70}
    #     modes = dict(zip(self.bmodes.copy(), self.mode_cols.copy()))
    #     for temp in temps:
    #         t_row = temps[temp]
    #         for mode in modes:
    #             m_col = modes[mode]
    #             for i in range(3):
    #                 voltage = ws.cell(row=t_row+i, column=2).value  ## hard-coded column 2 for voltages
    #                 min = ws.cell(row=t_row+i, column=m_col).value
    #                 max = ws.cell(row=t_row+i, column=m_col+1).value
    #                 min = round(float(min), 3)
    #                 max = round(float(max), 3)
    #                 voltage = float(voltage)
    #                 self.lim[temp][mode][voltage] = (min, max)


path = r"C:\Users\bruno\Desktop\P552 MCA PV AUX LIMITS.xlsx"
boards = ['B1', 'B2', 'B3', 'B4', 'B5', 'B6']
temps = [-40, 60, 85]
limits = Limits(path, "Sheet1", boards, temps)

print(limits.temp_rows)
print(limits.board_module_pairs)
print('MODULES:', limits.modules)
print('\nMode Row ===>', limits.mode_row, '\n')
print('MODULE MODES:', limits.module_modes)
print('BOARD MODES:', limits.board_modes)
print('Mode Columns:', limits.mode_cols)